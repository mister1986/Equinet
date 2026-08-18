"""
EQUINET Ad Scenarios — server-backed app.

The Excel file (equinet_database.xlsx) lives next to this script and is owned
entirely by the server. The browser never sees or touches the file directly —
it only talks to /api/data (read) and /api/save (write) over HTTP. Every
approval click or text edit in the page triggers an automatic POST, which
updates the .xlsx file on disk immediately. No file pickers, no downloads,
no user-facing file access at all.

Run:
    pip install flask openpyxl
    python server.py
Then open http://localhost:5000 in Chrome/Edge/Firefox/Safari — any browser
works now, since all the file I/O happens on the server, not in the browser.
"""
import json
import os
import threading
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
import openpyxl
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.resolve()
DB_PATH = BASE_DIR / "equinet_database.xlsx"
SEED_PATH = BASE_DIR / "seed_data.json"
SHEET_NAME = "ReviewData"

FIELD_LIST = ["h_on", "h_vo", "p_on", "p_vo", "s_on", "s_vo",
              "proof_on", "proof_vo", "b_on", "b_vo", "cta_vo"]
APPROVAL_FIELDS = ["mkApproved", "faezehApproved", "managerApproved"]
STATIC_FIELDS = ["row", "category", "funnel", "ceoApproved", "hook1",
                  "hook_url", "problem_url", "solution_url", "benefit_url"]

# All columns written to the spreadsheet, in order
COLUMNS = (STATIC_FIELDS + APPROVAL_FIELDS + FIELD_LIST +
           [f + "_editor" for f in FIELD_LIST])

_lock = threading.Lock()

app = Flask(__name__, static_folder=str(BASE_DIR / "static"))


def _load_seed():
    with open(SEED_PATH) as f:
        return json.load(f)


def _init_db_if_missing():
    if DB_PATH.exists():
        return
    seed = _load_seed()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(COLUMNS)
    for ad in seed:
        row = [ad.get(c, "") for c in COLUMNS]
        ws.append(row)
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(col) + 4))
    wb.save(DB_PATH)


def _read_db():
    _init_db_if_missing()
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[1]]
    ads = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, r))
        # normalize booleans (openpyxl may give 0/1/None)
        for bf in ["ceoApproved"] + APPROVAL_FIELDS:
            record[bf] = bool(record.get(bf))
        ads.append(record)
    return ads


def _write_field(row_num, field, value, editor=None):
    _init_db_if_missing()
    with _lock:
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb[SHEET_NAME]
        headers = [c.value for c in ws[1]]
        col_idx = headers.index(field) + 1
        row_idx = None
        row_col_idx = headers.index("row") + 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=row_col_idx).value == row_num:
                row_idx = r
                break
        if row_idx is None:
            raise ValueError(f"row {row_num} not found in database")
        ws.cell(row=row_idx, column=col_idx, value=value)
        if editor is not None and (field + "_editor") in headers:
            editor_col_idx = headers.index(field + "_editor") + 1
            ws.cell(row=row_idx, column=editor_col_idx, value=editor)
        wb.save(DB_PATH)


@app.route("/")
def index():
    return send_from_directory(str(BASE_DIR), "index.html")


@app.route("/api/data", methods=["GET"])
def get_data():
    return jsonify(_read_db())


@app.route("/api/save", methods=["POST"])
def save_field():
    payload = request.get_json(force=True)
    row = payload.get("row")
    field = payload.get("field")
    value = payload.get("value")
    editor = payload.get("editor")
    if row is None or field is None:
        return jsonify({"ok": False, "error": "row and field are required"}), 400
    try:
        _write_field(int(row), field, value, editor)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True})


if __name__ == "__main__":
    _init_db_if_missing()
    print(f"Database file: {DB_PATH}")
    print("Open http://localhost:5000 in your browser.")
    app.run(host="0.0.0.0", port=5000, debug=False)
